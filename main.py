import os
import re
import cv2
import math
import queue
import threading
import concurrent.futures
import numpy as np
import pandas as pd
import pytesseract
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
from openpyxl.drawing.image import Image as XLImage
# ====== TESSERACT ======
DEFAULT_TESSERACT = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
if os.path.exists(DEFAULT_TESSERACT):
    pytesseract.pytesseract.tesseract_cmd = DEFAULT_TESSERACT

# ====== QUEUE ======
log_queue = queue.Queue()

def log(msg):
    log_queue.put(msg)

# ====== IMAGE ======
def read_image_unicode(path):
    with open(path, 'rb') as f:
        data = np.frombuffer(f.read(), np.uint8)
    return cv2.imdecode(data, cv2.IMREAD_COLOR)

def preprocess(img):
    h, w = img.shape[:2]
    scale = max(1.5, 2500 / min(h, w))
    img = cv2.resize(img, None, fx=scale, fy=scale)

    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    _, th = cv2.threshold(gray, 180, 255, cv2.THRESH_BINARY_INV)

    kernel_h = cv2.getStructuringElement(cv2.MORPH_RECT, (50,1))
    kernel_v = cv2.getStructuringElement(cv2.MORPH_RECT, (1,50))

    th = cv2.subtract(th, cv2.morphologyEx(th, cv2.MORPH_OPEN, kernel_h))
    th = cv2.subtract(th, cv2.morphologyEx(th, cv2.MORPH_OPEN, kernel_v))

    return th

def preprocess_soft(img):
    h, w = img.shape[:2]
    scale = max(1.3, 2200 / min(h, w))
    img = cv2.resize(img, None, fx=scale, fy=scale)
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    _, th = cv2.threshold(gray, 200, 255, cv2.THRESH_BINARY_INV)
    return th

def split_4(img):
    h, w = img.shape[:2]
    return {
        "S21_amp": img[0:h//2, 0:w//2],
        "S21_gvz": img[0:h//2, w//2:w],
        "S11": img[h//2:h, 0:w//2],
        "S22": img[h//2:h, w//2:w]
    }

# ====== OCR ======
def ocr(img):
    return pytesseract.image_to_string(img, config="--psm 6 -l rus+eng").lower()

NUM_RE = re.compile(r"[-+]?\d+[.,]?\d*")
S11_RANGE = (1.0, 2.0)
S22_RANGE = (1.0, 2.0)
RX_ALLOWED_LEVELS = [40, 25, 10, -5, -20]
RX_TOLERANCE_DB = 4
TX_ALLOWED_LEVELS = [0, -15, -30]
TX_TOLERANCE_DB = 4
MAX_WORKERS = max(1, min(8, (os.cpu_count() or 1)))
RX_BANDS = {1: "1025-1525", 2: "975-1475", 3: "1076-1576"}
TX_BANDS = {1: "1000-1500", 2: "850-1350", 3: "1400-1900", 4: "1500-2000"}

def normalize_token(token):
    token = token.lower()
    translit = str.maketrans({
        "a": "а", "c": "с", "e": "е", "k": "к", "m": "м", "h": "н", "o": "о", "p": "р", "t": "т", "x": "х",
        "y": "у", "b": "в", "n": "п"
    })
    return token.translate(translit)

def parse_num(text):
    nums = NUM_RE.findall(text)
    if not nums:
        return None
    return float(nums[-1].replace(",", "."))

def parse_nums(text):
    return [float(n.replace(",", ".")) for n in NUM_RE.findall(text)]

def in_range(val, value_range):
    if value_range is None:
        return val is not None
    if val is None:
        return False
    lo, hi = value_range
    return lo <= val <= hi

def pick_number(candidates, value_range=None):
    if not candidates:
        return None
    if value_range is None:
        return candidates[-1]
    for n in reversed(candidates):
        if in_range(n, value_range):
            return n
    return None

def extract_from_lines(text, keys, value_range=None):
    lines = text.split("\n")
    norm_keys = [normalize_token(k) for k in keys]
    for i, line in enumerate(lines):
        norm_line = normalize_token(line)
        if any(k in norm_line for k in norm_keys):
            val = pick_number(parse_nums(line), value_range=value_range)
            if val is not None:
                return val
            if i + 1 < len(lines):
                val = pick_number(parse_nums(lines[i + 1]), value_range=value_range)
                if val is not None:
                    return val
    return None

def extract_from_layout(img, keys, value_range=None):
    data = pytesseract.image_to_data(
        img, config="--psm 6 -l rus+eng", output_type=pytesseract.Output.DICT
    )
    norm_keys = [normalize_token(k) for k in keys]
    tokens = []
    for i, raw in enumerate(data["text"]):
        text = raw.strip()
        conf = int(data["conf"][i]) if str(data["conf"][i]).lstrip("-").isdigit() else -1
        if not text or conf < 0:
            continue
        tokens.append({
            "text": text,
            "norm": normalize_token(text),
            "left": data["left"][i],
            "top": data["top"][i],
            "width": data["width"][i],
            "height": data["height"][i],
        })

    for tk in tokens:
        if not any(k in tk["norm"] for k in norm_keys):
            continue
        y = tk["top"]
        x_right = tk["left"] + tk["width"]
        candidates = []
        for other in tokens:
            if abs(other["top"] - y) > max(tk["height"], other["height"]) * 1.6:
                continue
            if other["left"] + other["width"] < x_right - 5:
                continue
            num = parse_num(other["text"])
            if num is None:
                continue
            if not in_range(num, value_range):
                continue
            dist = abs(other["left"] - x_right)
            candidates.append((dist, num))
        if candidates:
            candidates.sort(key=lambda item: item[0])
            return candidates[0][1]
    return None

def extract_metric(img, keys, value_range=None):
    text = ocr(img)
    val = extract_from_lines(text, keys, value_range=value_range)
    if val is not None:
        return val
    return extract_from_layout(img, keys, value_range=value_range)

def parse_image(path):
    img = read_image_unicode(path)
    zones = split_4(img)

    # Подготавливаем каждую зону отдельно, чтобы устойчиво работать
    # при "плавающей" области с текстовыми значениями.
    pre = {k: preprocess(v) for k, v in zones.items()}
    pre_soft = {k: preprocess_soft(v) for k, v in zones.items()}

    def metric_with_recovery(zone_name, keys, value_range):
        primary = extract_metric(pre[zone_name], keys, value_range=value_range)
        if primary is not None:
            return primary
        return extract_metric(pre_soft[zone_name], keys, value_range=value_range)

    return {
        "S21_amp_avg": metric_with_recovery("S21_amp", ["усил", "усип", "сред", "сред:", "cpea", "cped"], value_range=(-20, 45)),
        "S21_amp_uneven": metric_with_recovery("S21_amp", ["неравн", "неравн:", "нераен", "неревн"], value_range=(0.0, 6.5)),
        "S21_gvz_uneven": metric_with_recovery("S21_gvz", ["неравн", "неравн:", "нераен", "неревн"], value_range=(0, 20)),
        "S11_avg": metric_with_recovery("S11", ["сред", "сред:", "cpea", "cped"], value_range=S11_RANGE),
        "S22_avg": metric_with_recovery("S22", ["сред", "сред:", "cpea", "cped"], value_range=S22_RANGE)
    }

# ====== META ======
def detect_path(p):
    p=p.lower()
    if "tx" in p: return "Tx"
    if "rx" in p: return "Rx"
    return None

def detect_channel(n):
    n=n.lower()
    if "мейн" in n or "main" in n: return "Main"
    if "рез" in n or "reserve" in n: return "Reserve"
    return None

def detect_channel_no(name):
    name = name.lower()
    p = re.search(r"(?:канал\D*)(\d+)", name)
    if p:
        return int(p.group(1))
    p = re.search(r"\b(\d+)\s*канал", name)
    if p:
        return int(p.group(1))
    return None

def detect_att(n):
    nums = re.findall(r"\b(0|15|30)\b", n)
    if len(nums)==1: return f"ATT{nums[0]}"
    if len(nums)>=2: return f"{nums[0]}-{nums[1]}"
    return None

def detect_band(n):
    m=re.search(r"(\d{3,4})\D+(\d{3,4})", n)
    return f"{m.group(1)}-{m.group(2)}" if m else None

def metadata(path,file):
    p = detect_path(path)
    ch_no = detect_channel_no(file)
    mapped_band = RX_BANDS.get(ch_no) if p == "Rx" else TX_BANDS.get(ch_no) if p == "Tx" else None
    return {
        "Device": os.path.basename(os.path.dirname(os.path.dirname(path))),
        "Channel": detect_channel(file),
        "ChannelNo": ch_no,
        "Path": p,
        "Config": detect_att(file),
        "Band": mapped_band or detect_band(file)
    }

# ====== PHYSICS ======
def expected(cfg):
    if not cfg: return None
    nums=list(map(int,re.findall(r"\d+",cfg)))
    return 37 - sum(nums)

def conf(m,e):
    if m is None or e is None: return 0
    return math.exp(-abs(m-e)/10)

def conf_with_tol(measured, target, tol_db):
    if measured is None or target is None:
        return 0
    return max(0.0, 1 - abs(measured - target) / tol_db)

def conf_with_asym_tol(measured, target, minus_tol, plus_tol):
    if measured is None:
        return 0
    low = target + minus_tol
    high = target + plus_tol
    if low <= measured <= high:
        return 1.0
    if measured < low:
        d = low - measured
        span = max(0.001, abs(minus_tol))
        return max(0.0, 1 - d / span)
    d = measured - high
    span = max(0.001, abs(plus_tol))
    return max(0.0, 1 - d / span)

def conf_rx(m):
    if m is None:
        return 0
    nearest = min(RX_ALLOWED_LEVELS, key=lambda x: abs(m - x))
    return conf_with_tol(m, nearest, RX_TOLERANCE_DB)

def tx_target_from_config(cfg):
    if not cfg:
        return None
    vals = list(map(int, re.findall(r"\d+", cfg)))
    if not vals:
        return None
    total = -sum(vals)
    return min(TX_ALLOWED_LEVELS, key=lambda x: abs(x - total))

def expected_for_meta(meta):
    path = meta.get("Path")
    if path == "Rx":
        return None
    if path == "Tx":
        return tx_target_from_config(meta.get("Config"))
    return expected(meta.get("Config"))

def confidence_for_meta(meta, s21_amp):
    path = meta.get("Path")
    if path == "Rx":
        return conf_rx(s21_amp)
    if path == "Tx":
        target = tx_target_from_config(meta.get("Config"))
        return conf_with_tol(s21_amp, target, TX_TOLERANCE_DB)
    return conf(s21_amp, expected(meta.get("Config")))

def quality(vals, amp_conf):
    s11_ok = in_range(vals.get("S11_avg"), S11_RANGE)
    s22_ok = in_range(vals.get("S22_avg"), S22_RANGE)

    if amp_conf > 0.9 and s11_ok and s22_ok:
        return "OK"
    if amp_conf > 0.7 and (s11_ok or s22_ok):
        return "CHECK"
    return "BAD"

def process_file(path):
    vals = parse_image(path)
    meta = metadata(path, os.path.basename(path))
    exp = expected_for_meta(meta)
    c = confidence_for_meta(meta, vals["S21_amp_avg"])
    c_uneven = conf_with_asym_tol(vals.get("S21_amp_uneven"), 1.0, minus_tol=-0.8, plus_tol=5.0)
    q = quality(vals, c)
    return {
        **meta, **vals,
        "SourceImage": path,
        "Expected": exp,
        "Confidence": round(c, 3),
        "S21_amp_uneven_conf": round(c_uneven, 3),
        "Quality": q
    }

def write_report(rows, output_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    grouped = {}
    for row in rows:
        grouped.setdefault(row["Device"], []).append(row)

    headers = ["Preview", "Device", "Path", "Band", "Channel", "ChannelNo", "Config",
               "S21_amp_avg", "S21_amp_uneven", "S21_amp_uneven_conf", "S21_gvz_uneven",
               "S11_avg", "S22_avg", "Expected", "Confidence", "Quality"]

    for device, items in grouped.items():
        ws = wb.create_sheet(title=str(device)[:31] if device else "Unknown")
        ws.append([f"Device: {device}"])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        row_idx = 3

        for path_group in ("Rx", "Tx"):
            subset = [x for x in items if x.get("Path") == path_group]
            if not subset:
                continue
            ws.cell(row=row_idx, column=1, value=path_group)
            row_idx += 1
            subset.sort(key=lambda x: (str(x.get("Band") or ""), str(x.get("ChannelNo") or "")))
            band_groups = {}
            for it in subset:
                band_groups.setdefault(it.get("Band") or "Unknown", []).append(it)

            for band, band_items in band_groups.items():
                ws.cell(row=row_idx, column=1, value=f"Band {band}")
                row_idx += 1
                for c_idx, h in enumerate(headers, start=1):
                    ws.cell(row=row_idx, column=c_idx, value=h)
                row_idx += 1

                for it in band_items:
                    ws.row_dimensions[row_idx].height = 140
                    ws.cell(row=row_idx, column=2, value=it.get("Device"))
                    ws.cell(row=row_idx, column=3, value=it.get("Path"))
                    ws.cell(row=row_idx, column=4, value=it.get("Band"))
                    ws.cell(row=row_idx, column=5, value=it.get("Channel"))
                    ws.cell(row=row_idx, column=6, value=it.get("ChannelNo"))
                    ws.cell(row=row_idx, column=7, value=it.get("Config"))
                    ws.cell(row=row_idx, column=8, value=it.get("S21_amp_avg"))
                    ws.cell(row=row_idx, column=9, value=it.get("S21_amp_uneven"))
                    ws.cell(row=row_idx, column=10, value=it.get("S21_amp_uneven_conf"))
                    ws.cell(row=row_idx, column=11, value=it.get("S21_gvz_uneven"))
                    ws.cell(row=row_idx, column=12, value=it.get("S11_avg"))
                    ws.cell(row=row_idx, column=13, value=it.get("S22_avg"))
                    ws.cell(row=row_idx, column=14, value=it.get("Expected"))
                    ws.cell(row=row_idx, column=15, value=it.get("Confidence"))
                    ws.cell(row=row_idx, column=16, value=it.get("Quality"))

                    img_path = it.get("SourceImage")
                    if img_path and os.path.exists(img_path):
                        try:
                            img = XLImage(img_path)
                            img.width = 250
                            img.height = 140
                            ws.add_image(img, f"A{row_idx}")
                        except Exception:
                            ws.cell(row=row_idx, column=1, value=img_path)
                    row_idx += 1
                row_idx += 1

        for col in ("A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P"):
            ws.column_dimensions[col].width = 18 if col != "A" else 36

    wb.save(output_path)

# ====== WORKER ======
def worker(input_dir, output_dir, progress_var):
    rows=[]
    files=[]

    for r,_,f in os.walk(input_dir):
        for file in f:
            if file.lower().endswith((".png",".jpg",".jpeg")):
                files.append(os.path.join(r,file))

    total=len(files)
    if total == 0:
        log("Нет файлов изображений для обработки")
        return

    done = 0
    with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
        future_map = {ex.submit(process_file, path): path for path in files}

        for future in concurrent.futures.as_completed(future_map):
            path = future_map[future]
            try:
                rows.append(future.result())
                log(f"OK: {os.path.basename(path)}")
            except Exception as e:
                log(f"ERR: {path} -> {e}")
            done += 1
            progress_var.set(done/total*100)

    out=os.path.join(output_dir,"results.xlsx")
    write_report(rows, out)

    log("ГОТОВО")

# ====== GUI ======
class App:
    def __init__(self,root):
        self.root=root
        self.root.title("Парсер ВАЦ")

        self.input_dir=""
        self.output_dir=""

        tk.Button(root,text="Выбрать папку измерений",command=self.pick_input).pack(fill="x")
        tk.Button(root,text="Выбрать папку сохранения",command=self.pick_output).pack(fill="x")
        tk.Button(root,text="ЗАПУСТИТЬ",command=self.start).pack(fill="x")

        self.progress=ttk.Progressbar(root, length=300)
        self.progress.pack(fill="x")

        self.progress_var=tk.DoubleVar()
        self.progress.config(variable=self.progress_var)

        self.log=tk.Text(root,height=20)
        self.log.pack(fill="both",expand=True)

        self.update_log()

    def pick_input(self):
        self.input_dir=filedialog.askdirectory()
        log(f"Input: {self.input_dir}")

    def pick_output(self):
        self.output_dir=filedialog.askdirectory()
        log(f"Output: {self.output_dir}")

    def start(self):
        if not self.input_dir or not self.output_dir:
            messagebox.showerror("Ошибка","Выбери папки")
            return

        t=threading.Thread(target=worker,
                           args=(self.input_dir,self.output_dir,self.progress_var),
                           daemon=True)
        t.start()

    def update_log(self):
        while not log_queue.empty():
            msg=log_queue.get()
            self.log.insert(tk.END,msg+"\n")
            self.log.see(tk.END)
        self.root.after(100,self.update_log)

# ====== RUN ======
root=tk.Tk()
app=App(root)
root.mainloop()
